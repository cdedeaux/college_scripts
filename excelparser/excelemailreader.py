from openpyxl import *
import pandas as pd
wb = Workbook()
ws = wb.active
wb = load_workbook("snub.xlsx")
ws = wb.active
""
emails = []
"""
firstname = []
lastname = []
phone = []

"""
data = []
i = 0
for cell in ws["T"][155:234]:
    emails.append(str(cell.value))

"""
for cell in ws["B"][500:900]:
    firstname.append(str(cell.value))
for cell in ws["C"][500:900]:
    lastname.append(str(cell.value))
for cell in ws["D"][500:900]:
    phone.append(str(cell.value))
"""
df = pd.DataFrame(
    {
        "Email": emails
    }
)

df.to_csv("alumcontacts.csv", index=False)